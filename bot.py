import telebot
from requests.exceptions import ReadTimeout, ConnectionError
import sqlite3
from openpyxl import load_workbook, Workbook
import os
import logging
import sys
import time

# Инициализация логгера
def setup_logging(log_file):
    # Создаем объект логгера
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)  # Устанавливаем базовый уровень логирования на DEBUG

    # Создаем обработчик для записи в файл с указанием кодировки utf-8
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setLevel(logging.INFO)  # Устанавливаем уровень логирования INFO для файла

    # Создаем обработчик для вывода в консоль с указанием кодировки utf-8
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)  # Устанавливаем уровень логирования INFO для консоли

    # Создаем форматтер для сообщений
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

    # Устанавливаем форматтеры для обработчиков
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    # Добавляем обработчики к логгеру
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger
logging = setup_logging('bot.txt')

# Создаем экземпляр бота
TOKEN = "7037338775:AAFwN_SkX-MWVzfqE0OBfwEX3BYEeMm24yA"
bot = telebot.TeleBot(TOKEN)

# Функция для подключения к базе данных
def get_db_connection():
    conn = sqlite3.connect('tokens.db')
    cursor = conn.cursor()
    return conn, cursor

# Обработка команды /start и проверка токена
@bot.message_handler(commands=['start'])
def start(message):
    logging.info("[start] start mes")
    conn, cursor = get_db_connection()
    user_id = message.chat.id

    # Проверяем, привязан ли какой то токен к человеку
    cursor.execute("SELECT id, admin FROM user_token WHERE user_id_telegram=?", (user_id,))
    user_data = cursor.fetchone()
    conn.close()
    user_data = False
    if user_data:
        id, admin = user_data
        main_menu(message, admin, id)
    else:
        # Просим ввести токен
        msg = bot.send_message(message.chat.id, "Введите ваш токен:")
        bot.register_next_step_handler(msg, get_token)

# Проверяем токен от пользователя
def get_token(message):
    logging.info("[get_token]")
    conn, cursor = get_db_connection()
    chat_id = message.chat.id

    # Ищем токен
    cursor.execute("SELECT id, status, admin, user_id_telegram FROM user_token WHERE token=?", (message.text,))
    token_data = cursor.fetchone()
    conn.close()
    
    if token_data:
        id, status, admin, user_id_telegram = token_data
        if status and user_id_telegram != chat_id:
            # Если токена уже активирован
            bot.send_message(chat_id, "Токена уже занят!")
            # Просим ввести токен
            msg = bot.send_message(chat_id, "Введите ваш токен:")
            bot.register_next_step_handler(msg, get_token)
        else:
            main_menu(message, admin, id)
    else:
        # Если токена не существует
        bot.send_message(chat_id, "Такго токена не существует!")
        # Просим ввести токен
        msg = bot.send_message(chat_id, "Введите ваш токен:")
        bot.register_next_step_handler(msg, get_token)

# Главное меню
def main_menu(message, admin, id):
    logging.info(f"[main_menu] Проверка токена: admin - {admin}, id - {id}")
    user_id = message.chat.id

    if admin:
        # Привязываем токен и ставим статус
        conn, cursor = get_db_connection()
        cursor.execute("UPDATE user_token SET user_id_telegram = ?, status = ? WHERE id = ?", (user_id, True, id,))
        conn.commit()
        conn.close()
        # Откраываем админ панель
        admin_panel(message)
    else:
        # Привязываем токен и ставим статус
        conn, cursor = get_db_connection()
        cursor.execute("UPDATE user_token SET user_id_telegram = ?, status = ? WHERE id = ?", (user_id, True, id,))
        conn.commit()
        conn.close()
        # Открывает пользовательнсую панель
        user_panel(message) 



# Панель администратора
def admin_panel(message):
    logging.info("[admin_panel]")
    # Выводим кнопки "Новый токен" и "Список токенов"
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn_new_token = telebot.types.KeyboardButton('Новый токен')
    btn_token_list = telebot.types.KeyboardButton('Список токенов')
    markup.add(btn_new_token, btn_token_list)
    bot.send_message(message.chat.id, "Выберите действие:", reply_markup=markup)


# Обработка нажатия кнопки "Новый токен"
@bot.message_handler(func=lambda message: message.text == 'Новый токен')
def new_token(message):
    logging.info("[new_token]")
    conn, cursor = get_db_connection()
    user_id = message.chat.id

    # Проверяем, привязан ли какой то токен к человеку
    cursor.execute("SELECT id, admin FROM user_token WHERE user_id_telegram=?", (user_id,))
    user_data = cursor.fetchone()
    conn.close()

    if not user_data:
        markup = telebot.types.ReplyKeyboardRemove()
        bot.send_message(message.chat.id, "Ваш токен был удален.", reply_markup=markup)
        start(message)
    else:
        msg = bot.send_message(message.chat.id, "Напишите для чего этот токен, чтобы потом не забыть:")
        bot.register_next_step_handler(msg, process_token_description)

# Обработка введенной информации (описания токена)
def process_token_description(message):
    logging.info("[process_token_description]")
    token_description = message.text  # Описание токена, введенное пользователем
    msg = bot.send_message(message.chat.id, "Теперь напишите сам токен:")
    bot.register_next_step_handler(msg, process_token_input, token_description)

# Обработка токена и запись данных в базу
def process_token_input(message, token_description):
    token_value = message.text  # Сам токен, введенный пользователем
    chat_id = message.chat.id
    logging.info(f"[process_token_input] Сохранение токена: token - {token_value}, Описание - {token_description}")

    # Добавляем данные в базу данных
    conn, cursor = get_db_connection()
    try:
        cursor.execute("""
            INSERT INTO user_token (token, status, admin, name, user_id_telegram)
            VALUES (?, ?, ?, ?, ?)""",
            (token_value, False, False, token_description, None))
        conn.commit()
        conn.close()

        bot.send_message(chat_id, f"Создан токен: ```{token_value}```Описание: {token_description}", parse_mode='Markdown')
    except sqlite3.IntegrityError:
        bot.send_message(chat_id, "Токен с таким названием уже существует!")
        process_token_input(message, token_description)


# Обработка нажатия кнопки "Список токенов"
@bot.message_handler(func=lambda message: message.text == 'Список токенов')
def list_token(message, page=1):
    logging.info(f"[list_token] Ввод страницы: {page}")
    conn, cursor = get_db_connection()
    user_id = message.chat.id

    # Проверяем, привязан ли какой то токен к человеку
    cursor.execute("SELECT id, admin FROM user_token WHERE user_id_telegram=?", (user_id,))
    user_data = cursor.fetchone()

    if not user_data:
        conn.close()
        markup = telebot.types.ReplyKeyboardRemove()
        bot.send_message(message.chat.id, "Ваш токен был удален.", reply_markup=markup)
        start(message)
    else:
        # Получаем все токены из базы данных
        cursor.execute("SELECT id, token FROM user_token")
        tokens = cursor.fetchall()
        conn.close()

        # Определяем количество токенов на одной странице
        tokens_per_page = 10
        total_pages = (len(tokens) - 1) // tokens_per_page + 1

        # Получаем токены для текущей страницы
        start_idx = (page - 1) * tokens_per_page
        end_idx = start_idx + tokens_per_page
        tokens_on_page = tokens[start_idx:end_idx]

        # Формируем кнопки с токенами (по id записи)
        markup = telebot.types.InlineKeyboardMarkup(row_width=1)
        for token in tokens_on_page:
            token_id = token[0]  # id записи
            token_value = token[1]  # токен
            btn = telebot.types.InlineKeyboardButton(f"{token_value}", callback_data=f"token_{token_id}")
            markup.add(btn)

        # Добавляем кнопки навигации
        if page > 1:
            btn_prev = telebot.types.InlineKeyboardButton('⬅️ Назад', callback_data=f'prev_page_{page - 1}')
            markup.add(btn_prev)
        
        if page < total_pages:
            btn_next = telebot.types.InlineKeyboardButton('➡️ Вперед', callback_data=f'next_page_{page + 1}')
            markup.add(btn_next)

        # Отправляем сообщение с токенами и кнопками навигации
        bot.send_message(message.chat.id, f"Страница {page} из {total_pages}", reply_markup=markup)

# Обработчик для кнопок навигации (Вперед и Назад)
@bot.callback_query_handler(func=lambda call: call.data.startswith('prev_page_') or call.data.startswith('next_page_'))
def handle_page_navigation(call):
    page = int(call.data.split('_')[-1])
    logging.info(f"[handle_page_navigation] Переход на страницу: {page}")
    list_token(call.message, page=page)

# Обработчик для выбора токена (при нажатии на кнопку с id записи)
@bot.callback_query_handler(func=lambda call: call.data.startswith('token_'))
def handle_token_selection(call):
    logging.info("[handle_token_selection] Открываем данные токена")
    selected_token_id = call.data.split('_')[1]  # Получаем id записи
    
    # Подключаемся к базе данных
    conn, cursor = get_db_connection()
    
    # Извлекаем всю информацию о токене по id записи
    cursor.execute("SELECT token, status, admin, name FROM user_token WHERE id = ?", (selected_token_id,))
    token_data = cursor.fetchone()
    conn.close()

    if token_data:
        token_value, status, admin, name = token_data
        status_text = "Активирован" if status else "Не активирован"
        admin_text = "Да" if admin else "Нет"
        
        # Формируем сообщение с полной информацией о токене
        token_info = (f"Токен: ```{token_value}```\n"
                      f"Статус: {status_text}\n"
                      f"admin: {admin_text}\n"
                      f"Описание: {name}")
        
        # Создаем инлайн-кнопку "Удалить"
        markup = telebot.types.InlineKeyboardMarkup()
        btn_delete = telebot.types.InlineKeyboardButton("Удалить", callback_data=f'delete_{selected_token_id}')
        markup.add(btn_delete)

        # Отправляем сообщение с информацией о токене и кнопкой
        bot.send_message(call.message.chat.id, token_info, reply_markup=markup, parse_mode="Markdown")
    else:
        bot.send_message(call.message.chat.id, "Информация о токене не найдена.")

# Обработчик для удаления токена по id записи
@bot.callback_query_handler(func=lambda call: call.data.startswith('delete_'))
def handle_token_delete(call):
    logging.info("[handle_token_delete] Удаляем токен")
    selected_token_id = call.data.split('_')[1]  # Получаем id записи, который нужно удалить
    
    # Подключаемся к базе данных
    conn, cursor = get_db_connection()
    
    # Удаляем токен по id записи
    cursor.execute("DELETE FROM user_token WHERE id = ?", (selected_token_id,))
    conn.commit()
    conn.close()

    # Отправляем сообщение о том, что токен был удален
    bot.send_message(call.message.chat.id, f"Токен с ID {selected_token_id} был удален.", parse_mode="Markdown")



# user panel
def user_panel(message):
    logging.info("[user_panel]")
    # Выводим кнопки "Создать заказ" и "Посмотреть заказы"
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn_create_order = telebot.types.KeyboardButton('Создать заказ')
    btn_view_orders = telebot.types.KeyboardButton('Посмотреть заказы')
    markup.add(btn_create_order, btn_view_orders)
    bot.send_message(message.chat.id, "Выберите действие:", reply_markup=markup)


# Обработка нажатия кнопки "Создать заказ"
@bot.message_handler(func=lambda message: message.text == 'Создать заказ')
def new_order(message):
    logging.info("[new_order] Начало нового заказа")
    conn, cursor = get_db_connection()
    user_id = message.chat.id

    # Проверяем, привязан ли какой то токен к человеку
    cursor.execute("SELECT id, admin FROM user_token WHERE user_id_telegram=?", (user_id,))
    user_data = cursor.fetchone()
    conn.close()

    if not user_data:
        markup = telebot.types.ReplyKeyboardRemove()
        bot.send_message(message.chat.id, "Ваш токен был удален.", reply_markup=markup)
        start(message)
    else:
        # Спрашиваем у пользователя паспорт объекта
        msg = bot.send_message(message.chat.id, "Введите паспорт объекта:")
        bot.register_next_step_handler(msg, process_passport_input)

# Обработчик для получения паспорта объекта
def process_passport_input(message):
    user_passport = message.text  # Сохраняем паспорт объекта
    logging.info(f"[process_passport_input] Сохраняем паспорт: {user_passport}")

    # Запрашиваем текст или аудио файл
    msg = bot.send_message(message.chat.id, "Напишите данные заказа, отправьте файл или аудио:")
    bot.register_next_step_handler(msg, process_file_input, user_passport)

# Обработчик для получения текста заказа loop
def process_passport_input_loop(message, user_passport):
    logging.info("[process_passport_input_loop] Ждем текст заказа")
    # Запрашиваем текст или аудио файл
    msg = bot.send_message(message.chat.id, "Напишите данные заказа, отправьте файл или аудио:")
    bot.register_next_step_handler(msg, process_file_input, user_passport)

# Обработчик для начала ввода заказа
def process_file_input(message, user_passport, user_data={'text': "",'files': []}):
    logging.info(f"[process_file_input] Получаем текст заказа и работаем по кругу до исключения: {user_data}")
    user_id = message.chat.id

    if message.text == 'Закончить':
        finish_input(message, user_passport, user_data)
    else:
        # Проверяем, что было отправлено пользователем
        if message.content_type == 'text':
            user_data['text'] += message.text + "\n"
        elif message.content_type == 'document':
            file_info = bot.get_file(message.document.file_id)
            file_link = f"https://api.telegram.org/file/bot{TOKEN}/{file_info.file_path}"
            user_data['files'].append(file_link)
        elif message.content_type == 'audio':
            file_info = bot.get_file(message.audio.file_id)
            file_link = f"https://api.telegram.org/file/bot{TOKEN}/{file_info.file_path}"
            user_data['files'].append(file_link)
        elif message.content_type == 'photo':
            file_info = bot.get_file(message.photo[-1].file_id)  # Берем последнее фото в списке (наивысшее качество)
            file_link = f"https://api.telegram.org/file/bot{TOKEN}/{file_info.file_path}"
            user_data['files'].append(file_link)

        # Отправляем сообщение с кнопкой "Закончить"
        markup = telebot.types.ReplyKeyboardMarkup(one_time_keyboard=True)
        markup.add("Закончить")
        msg = bot.send_message(user_id, "Сообщение получено. Нажмите 'Закончить', когда завершите ввод.", reply_markup=markup)
        bot.register_next_step_handler(msg, process_file_input, user_passport, user_data)

# Обработчик для завершения ввода и сохранения данных
def finish_input(message, user_passport, user_data):
    logging.info(f"[finish_input] Обработка исключения: паспорт - {user_passport}, данные заказа - {user_data}")
    user_id = message.chat.id
    
    # Получаем данные пользователя
    text_content = user_data['text']
    file_links = ", ".join(user_data['files']) if user_data['files'] else "Нет файлов"

    # Записываем заказ в базу данных
    save_order_to_db(user_id, text_content, file_links, user_passport)

    # Формируем сообщение с данными заказа для пользователя
    order_summary = f"""
Ваш заказ был успешно сохранен.
    
Паспорт объекта: {user_passport}
Текст заказа:\n{text_content if text_content else "Текст отсутствует"}Ссылки на файлы:\n{file_links}
    """
    
    # Отправляем пользователю сообщение с данными заказа
    bot.send_message(user_id, order_summary)
    
    # Очищаем временные данные пользователя
    del user_data

    user_panel(message)

# Функция для записи заказа в базу данных
def save_order_to_db(user_id, text_content, file_content, user_passport):
    logging.info("[save_order_to_db] Сохраняем в бд заказ")
    conn, cursor = get_db_connection()

    # Добавляем новый заказ в таблицу orders
    cursor.execute("""
        INSERT INTO orders (user_id_telegram, text, file, status, pass)
        VALUES (?, ?, ?, ?, ?)
    """, (user_id, text_content, file_content, 'Ожидание', user_passport))

    conn.commit()
    conn.close()

    # Сохранение данных в Excel файл
    save_order_to_excel(text_content, file_content, user_passport)

# Функция для сохранения данных в Excel файл
def save_order_to_excel(text_content, file_content, user_passport):
    logging.info("[save_order_to_excel] Сохраняем в эксель заказ")
    file_name = 'orders.xlsx'

    # Проверяем, существует ли файл Excel
    if os.path.exists(file_name):
        # Загружаем существующий файл
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        # Создаем новый Excel файл и добавляем заголовки
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Паспорт обьекта", "Текст заказа", "Файлы", "Статус"])

    # Добавляем новую строку данных
    sheet.append([user_passport, text_content, file_content, "Ожидание"])

    # Сохраняем изменения в файл
    workbook.save(file_name)


# Обработка нажатия кнопки "Посмотреть заказы"
@bot.message_handler(func=lambda message: message.text == 'Посмотреть заказы')
def list_user_order(message, page=1):
    logging.info("[list_user_order] Выводим список заказов")
    conn, cursor = get_db_connection()
    user_id = message.chat.id

    # Проверяем, привязан ли какой то токен к человеку
    cursor.execute("SELECT id, admin FROM user_token WHERE user_id_telegram=?", (user_id,))
    user_data = cursor.fetchone()

    if not user_data:
        conn.close()
        markup = telebot.types.ReplyKeyboardRemove()
        bot.send_message(message.chat.id, "Ваш токен был удален.", reply_markup=markup)
        start(message)
    else:
        # Извлекаем заказы пользователя из базы данных
        cursor.execute("SELECT id, pass FROM orders WHERE user_id_telegram = ?", (user_id,))
        orders = cursor.fetchall()
        conn.close()

        # Если нет заказов, сообщаем об этом
        if not orders:
            bot.send_message(message.chat.id, "У вас нет заказов.")
            return

        # Определяем количество заказов на странице
        orders_per_page = 10
        start = (page - 1) * orders_per_page
        end = start + orders_per_page
        paginated_orders = orders[start:end]

        # Создаем инлайн-кнопки для заказов
        markup = telebot.types.InlineKeyboardMarkup()
        for order in paginated_orders:
            order_id, pass_info = order
            btn_order = telebot.types.InlineKeyboardButton(
                text=f"Заказ #{order_id} - {pass_info}",
                callback_data=f'order_{order_id}'
            )
            markup.add(btn_order)

        # Добавляем навигацию по страницам
        navigation_buttons = []
        if page > 1:
            prev_btn = telebot.types.InlineKeyboardButton("⬅️", callback_data=f'orders_page_{page-1}')
            navigation_buttons.append(prev_btn)
        if len(orders) > end:
            next_btn = telebot.types.InlineKeyboardButton("➡️", callback_data=f'orders_page_{page+1}')
            navigation_buttons.append(next_btn)

        if navigation_buttons:
            markup.add(*navigation_buttons)

        # Отправляем сообщение с кнопками
        bot.send_message(message.chat.id, f"Страниц {page}\\{end}\nВаши заказы:", reply_markup=markup)

# Обработчик навигации по страницам заказов
@bot.callback_query_handler(func=lambda call: call.data.startswith('orders_page_'))
def handle_orders_pagination(call):
    # Получаем номер страницы
    page = int(call.data.split('_')[2])
    logging.info(f"[handle_orders_pagination] переход на страницу: {page}")
    list_user_order(call.message, page)

# Обработчик для выбора заказа (при нажатии на кнопку с заказом)
@bot.callback_query_handler(func=lambda call: call.data.startswith('order_'))
def handle_order_selection(call):
    order_id = call.data.split('_')[1]  # Получаем ID заказа
    logging.info(f"[handle_orders_pagination] Выводим данные заказа: id - {order_id}")
    
    # Подключаемся к базе данных
    conn, cursor = get_db_connection()
    
    # Извлекаем информацию о заказе
    cursor.execute("SELECT pass, text, file, status FROM orders WHERE id = ?", (order_id,))
    order_data = cursor.fetchone()
    conn.close()

    if order_data:
        pass_info, text_content, file, status = order_data
        
        # Формируем и отправляем сообщение с информацией о заказе
        order_info = (f"Паспорт объекта: {pass_info}\n"
                      f"Текст:\n{text_content or 'Текст не добавлен'}"
                      f"Файлы:\n{file or 'Файл не добавлен'}\n"
                      f"Статус: {status}")
        bot.send_message(call.message.chat.id, order_info)
    else:
        bot.send_message(call.message.chat.id, "Информация о заказе не найдена.")



# Запускаем бота
while True:
    try:
        logging.info("[start_bot] Starting bot polling")
        bot.polling(none_stop=True)
    except ReadTimeout:
        logging.warning("ReadTimeout error occurred. Retrying in 5 seconds...")
        time.sleep(5)
    except ConnectionError:
        logging.warning("ConnectionError occurred. Retrying in 5 seconds...")
        time.sleep(5)
    except Exception as e:
        logging.warning(f"An unexpected error occurred: {e}. Retrying in 5 seconds...")
        time.sleep(5)